#! /usr/bin/env python3.10
# messages.py - sends text and email messages

import smtplib
import ssl
import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill
from email.message import EmailMessage
from email.utils import formataddr
from mimetypes import guess_type
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from jinja2 import Environment, FileSystemLoader
from typing import Optional, Any
from dotenv import load_dotenv

load_dotenv()

# -- Constants ------------------------------------------
HEADER = ["Name", "Phone", "Email Address"]
HEADER_COLOR = "0056B3"
HEADER_FONT = "FFFFFF"


def render_template(template_file: str, **kwargs: Any) -> str:
    # Load templates from the current directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    templates_dir = os.path.join(base_dir, "templates")
    env = Environment(loader=FileSystemLoader(templates_dir))
    template = env.get_template(template_file)
    return template.render(**kwargs)


def contacts_mailinglist(cursor: sqlite3.Cursor) -> list[dict]:
    results = []

    cursor.execute("SELECT * FROM contacts")
    contacts = cursor.fetchall()

    if not contacts:
        print("No contacts in database.")
        return []

    for contact in contacts:
        results.append(
            {
                "id": contact[0],
                "name": contact[1],
                "phone": contact[2],
                "email_address": contact[3],
                "date_added": contact[4],
            }
        )
    return results


def open_or_create_email_database(db_name: str) -> tuple[sqlite3.Connection, sqlite3.Cursor]:
    # your database code here
    """CREATE the database and table"""
    # Connect to database (create file if it doesn't exist)
    conn = sqlite3.connect(db_name, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    cursor = conn.cursor()

    # Create the table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT NOT NULL,
            email_address TEXT NOT NULL UNIQUE,
            date_added DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Create answer table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS contact_answers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            answer_1 TEXT NOT NULL,
            answer_2 TEXT NOT NULL,
            answer_3 TEXT NOT NULL
        )
    """)

    conn.commit()
    print("Database and table created successfully!")
    return conn, cursor


def close_database(conn: sqlite3.Connection):
    conn.close()
    print("Database connection closed.")


def add_contacts(
    conn: sqlite3.Connection, cursor: sqlite3.Cursor, name: str, phone: str, email_address: str, date_added: str
) -> None:
    """Add a contact to the database"""
    try:
        cursor.execute(
            """
            INSERT INTO contacts(name, phone, email_address, date_added)
            VALUES (?, ?, ?, ?)
        """,
            (name, phone, email_address, date_added),
        )
        conn.commit()

        print(f"Added : {name} {email_address}")
    except sqlite3.IntegrityError:
        print(f"❌Error: Contact '{name}' '{email_address}' already exists!")


def add_multiple_contacts(conn: sqlite3.Connection, cursor: sqlite3.Cursor, contacts_list: list) -> None:
    """Add multiple contacts to the database"""
    for name, phone, email_address, date_added in contacts_list:
        add_contacts(conn, cursor, name, phone, email_address, date_added)


def delete_contacts(conn: sqlite3.Connection, cursor: sqlite3.Cursor, name: str) -> None:
    """Delete contact by NAME"""
    cursor.execute("SELECT * FROM contacts WHERE name = ?", (name,))
    result = cursor.fetchone()

    if result:
        cursor.execute("DELETE FROM contacts WHERE id = ?", (result[0],))
        conn.commit()
        print(f"Deleted: {result[1]} | {result[3]}")
    else:
        print(f"No contact found matching '{name}'")


def view_all_contacts(cursor: sqlite3.Cursor) -> None:
    """Display all contacts in the database"""
    cursor.execute("SELECT * FROM contacts")
    contacts = cursor.fetchall()

    if not contacts:
        print("No contacts in database.")
        return

    print("\n" + "=" * 80)
    print("ALL CONTACTS IN DATABASE")
    print("-" * 80)
    for contact in contacts:
        print(f"\nID: {contact[0]}")
        print(f"Name: {contact[1]}")
        print(f"Phone: {contact[2]}")
        print(f"Email Address: {contact[3]}")
        print(f"Date Added: {contact[4]}")
        print("-" * 80)


def search_contacts(cursor: sqlite3.Cursor, search_term: str) -> None:
    """Search contacts by name email_address or phone"""
    cursor.execute(
        """
        SELECT * FROM contacts
        WHERE name LIKE ? OR phone LIKE ? OR email_address LIKE ?
    """,
        (f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"),
    )

    result = cursor.fetchall()

    if not result:
        print(f"No contacts found matching '{search_term}'")
        return

    print(f"\n🔍 Found {len(result)} contact(s):")
    for contact in result:
        print(f"\n  • {contact[1]} {contact[3]}")
        print(f"    {contact[4]}")


def update_contact(
    conn: sqlite3.Connection, cursor: sqlite3.Cursor, contact_id: int, name: str, field: str, new_value: str
) -> None:
    """Update a specific field of a contact"""
    valid_fields = ["name", "phone", "email_address", "date_added"]

    if field not in valid_fields:
        print(f"❌ Invalid field. Choose from: {valid_fields}")
        return

    cursor.execute(
        f"""
        UPDATE contacts
        SET {field} = ?
        WHERE id = ?
    """,
        (new_value, contact_id),
    )

    conn.commit()
    print(f"✅ Updated {field} for contact ID {contact_id}")


def open_or_create_email_workbook(filename: str):
    # your openpyxl code here
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        print(f"Opened existing workbook: {filename}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Contacts_1"
        write_header(ws)
        wb.save(filename)
        print(f"Created new workbook: {filename}")
    return wb


# Header writer
def write_header(sheet):
    sheet.append(HEADER)
    for col, cell in enumerate(sheet[1], start=1):
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")


class Messages:
    def __init__(self, name: str, phone: str):
        self.name = name
        self.phone = phone

    def send(self):
        raise NotImplementedError("Subclass must implement send()")

    def __str__(self) -> str:
        return f"{self.name} | {self.phone}"


# Email subclass
class EmailMessages(Messages):
    def __init__(
        self,
        name: str,
        phone: str,
        email_address: str,
        subject: str,
        body: str,
        smtp_server: str,
        smtp_port: int,
        username: str,
        password: str,
    ) -> None:

        super().__init__(name, phone)  # inherits name and phone from Message
        self.email = email_address
        self.subject = subject
        self.body = body
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.username = username
        self.password = password

    def send(self):
        try:
            # Build the email
            msg = MIMEMultipart("alternative")
            msg["Subject"] = self.subject  # <- self. instead of parameter
            msg["From"] = self.username  # <- self. instead of parameter
            msg["To"] = self.email  # <- self. instead of parameter

            # Attach the HTML body
            msg.attach(MIMEText(self.body, "html"))

            # Connect and send
            with smtplib.SMTP_SSL(self.smtp_server, self.smtp_port) as server:
                server.login(self.username, self.password)
                server.sendmail(self.username, self.email, msg.as_string())
                print(f"Email sent successfully to {self.email}")

        except smtplib.SMTPException as e:
            print(f"Failed to send email: {e}")


# AutomateEmail subclass
class AutomateEmailMessage(Messages):
    def __init__(
        self,
        name: str,
        phone: str,
        sender_email: str,
        sender_name: str,
        password: str,
        email_address: str,
        email_body: str,
        email_subject: str,
        is_html: bool,
        attachments: Optional[list[str]] = None,
    ) -> None:

        super().__init__(name, phone)  # inherits name and phone from Message
        self.sender_email = sender_email
        self.sender_name = sender_name
        self.password = password
        self.email_address = email_address
        self.email_body = email_body
        self.email_subject = email_subject
        self.is_html = is_html
        self.attachments = attachments
        self.smtp_port = 587
        self.smtp_server = "smtp.gmail.com"
        self.questions = [
            "How satisfied are you with our service?",
            "Would you recommend us to a friend?",
            "How can we improve?",
        ]

    def send(self):
        html_body = render_template(
            "automate_email_message_template4.html",
            subject="Your Learning Dreams Visit Confirmation",
            company_name="Learning Dreams",
            company_address="1091 Creekwood Trail, Burton, MI 48509",
            recipient_name=self.name,
            sender_name="Customer Service",
            body="Thank you for your visit. Here is your summary:",
            items=["Test 1 — Evaluation"],  # "Item 2 — Gadget", "Item 3 — Doohickey"],
            questions=self.questions,
            unsubscribe_link="http://localhost:5000/unsubscribe",
            feedback_link="http://localhost:5000/contacts",
            attachments=self.attachments,
        )

        msg = EmailMessage()
        msg["Subject"] = self.email_subject
        msg["From"] = formataddr((self.sender_name, self.sender_email))

        # Support both plain text and HTML emails
        if self.is_html:
            msg.set_content("Please enable HTML to view this email.")
            msg.add_alternative(html_body, subtype="html")
        else:
            msg.set_content(self.email_body)

        # Add attachments if provided
        if self.attachments:
            for file_path in self.attachments:
                try:
                    with open(file_path, "rb") as file:
                        file_data = file.read()
                        file_name = os.path.basename(file_path)
                        mime_type, _ = guess_type(file_path)
                        if mime_type:
                            mime_main, mime_subtype = mime_type.split("/")
                        else:
                            mime_main, mime_subtype = "application", "octet-stream"

                        msg.add_attachment(file_data, maintype=mime_main, subtype=mime_subtype, filename=file_name)
                        print(f"Attached file: {file_name}")
                except Exception as e:
                    print(f"Failed to attach {file_path}: {e}")

        # Adding ssl layer of security
        ssl_context = ssl.create_default_context()

        try:
            # Creating smtp server
            print("Connecting to Server....")
            my_server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            my_server.starttls(context=ssl_context)

            # Login to smtp server
            my_server.login(self.sender_email, self.password)
            print("Connected to server!")

            # Sending email
            print(f"Sending email from: {self.sender_email}")
            print("********************************************")
            # for receiver:
            msg["To"] = self.receiver
            print(f"Sending email to: {self.receiver}")
            my_server.send_message(msg)
            print(f"....\nSuccessfully sent to: {self.receiver}")
            print("********************************************")
            del msg["To"]
        except Exception as e:
            print(f"ERROR: {e}")
        finally:
            my_server.quit()


# Text subclass
class TextMessage(Messages):
    def __init__(self, name: str, phone: str, body: str) -> None:
        super().__init__(name, phone)  # inherits name and phone from Message
        self.body = body

    def send(self):
        # Your iMessage sending code here
        print(f"Sending text to {self.phone}")


if __name__ == "__main__":
    attachments = [os, environ.get("ATTACHMENT_PATH") or ""]
    FILENAME = os.environ.get("DB_FILENAME") or "contacts.db"
    conn, cursor = open_or_create_email_database(FILENAME)
    add_contacts(conn, cursor, "Your Name", "+1234567890", "your@email.com", "MM/DD/YYYY")
    add_contacts(conn, cursor, "Your Name", "+1234567890", "your@email.com", "MM/DD/YYYY")
    mailing_list = contacts_mailinglist(cursor)

    if not mailing_list:
        print("No contacts found.")
    else:
        for i in range(len(mailing_list)):
            html_body = render_template(
                "automate_email_message_template4.html",
                subject="Your Subject",
                company_name="Your Company Name",
                company_address="Your Company Address",
                recipient_name=mailing_list[i]["name"],
                sender_name="Your Name",
                body="We appreciate your continued support. Please click the link below to submit your feedback: https://RussellCarsonJr.pythonanywhere.com/contacts",
                questions=[
                    "How satisfied are you with our service?",
                    "Would you recommend us to a friend?",
                    "How can we improve?",
                ],
                items=[],
                feedback_link="http://localhost:5000/contacts",
                unsubscribe_link="http://localhost:5000/unsubscribe",
            )

            message = AutomateEmailMessage(
                name=mailing_list[i]["name"],
                phone=mailing_list[i]["phone"],
                sender_email=os.environ.get("SENDER_EMAIL") or "",
                sender_name="Your Name",
                password=os.environ.get("EMAIL_PASSWORD") or "",
                receiver=mailing_list[i]["email_address"],
                email_body=html_body,
                email_subject="Your Subject",
                is_html=True,
                attachments=attachments,
            )
            message.send()
